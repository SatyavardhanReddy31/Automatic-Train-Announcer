import openpyxl
import pyttsx3
import time
import playsound
i = int(input('ENTER THE TRAIN S.NO. AS PREFERRED IN FILE: '))
j = int(input('ENTER THE PLATFORM NUMBER TRAIN ENTERS: '))
speaker = pyttsx3.init()
voices = speaker.getProperty('voices')
speaker.setProperty('voice',"HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\Tokens\TTS_MS_EN-US_ZIRA_11.0")
wb = openpyxl.load_workbook("names.xlsx")
sheet = wb.sheetnames
print(sheet)
sh = wb["train"]
def f1():
    playsound.playsound("t101.mp3")
    time.sleep(1)
    speaker.setProperty("rate",100)
    speaker.say(sh.cell(i+1,2).value)
    speaker.runAndWait()
def f2():
    speaker = pyttsx3.init()
    speaker.setProperty("rate", 100)
    speaker.say(sh.cell(i+1,3).value)
    speaker.runAndWait()
    playsound.playsound("t102.mp3")
    speaker = pyttsx3.init()
    speaker.setProperty("rate", 100)
    speaker.say(sh.cell(i+1,4).value)
    speaker.runAndWait()
def f3():
    speaker = pyttsx3.init()
    speaker.setProperty("rate", 100)
    speaker.say(sh.cell(i + 1, 5).value)
    speaker.runAndWait()
    playsound.playsound("t103.mp3")
    speaker = pyttsx3.init()
    speaker.setProperty("rate", 100)
    speaker.say(sh.cell(i+1,1).value)
    speaker.runAndWait()
    playsound.playsound("train104.mp3")
def f4():
    speaker = pyttsx3.init()
    speaker.setProperty("rate", 100)
    speaker.say(j)
    speaker.runAndWait()
    playsound.playsound("train105.mp3")
    quit()
def call():
    print("ANNOUNCING.........")
    f1()
    f2()
    f3()
    f4()
    quit()
call()